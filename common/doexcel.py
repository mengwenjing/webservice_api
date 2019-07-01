from openpyxl import load_workbook

class Case:

    def __init__(self):
        self.case_id = None
        self.title = None
        self.url = None
        self.data = None
        self.expected = None
        self.actual = None
        self.result = None
        self.sql = None

class DoExcel:
    """对Excel里的数据进行读写操作"""
    def __init__(self, filename, sheet):
        try:
            self.filename = filename
            self.workbook = load_workbook(filename)
            self.sheet = self.workbook[sheet]
        except Exception as e:
            print('找不到该文件，错误是：{}'.format(e))

    def read_data(self):
        big_data = []#定义一个空列表，用来存放读取出来的所有数据
        for i in range(2, self.sheet.max_row+1):
            data = Case()
            data.case_id = self.sheet.cell(i,1).value
            data.title = self.sheet.cell(i,2).value
            data.url = self.sheet.cell(i,3).value
            data.data = self.sheet.cell(i,4).value
            data.expected = self.sheet.cell(i,5).value
            data.sql = self.sheet.cell(i,8).value
            big_data.append(data)
        self.workbook.close()
        return big_data

    def write_data(self,row,actual,result):#指定写入内容到底6，7列
        self.sheet.cell(row,6).value = actual
        self.sheet.cell(row,7).value = result
        self.workbook.save(self.filename)
        self.workbook.close()

if __name__ == '__main__':
    from common.contants import case_file
    r = DoExcel(case_file,'sendMCode').read_data()
    for i in r:
        print(i.case_id)


